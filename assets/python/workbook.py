#!/usr/bin/env python3
import os
import csv
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
#imports the modules needed to get this thing working


# Create the function scan_book_capital
def scan_book_capital(fp):
    """
    Read only the 'Book Capital' sheet in read_only mode
    and extract all K1 ID → Ending Capital Balance pairs,
    plus the Fund (parent folder name).
    """

    #creates an empty list to later store records
    records = []


    # derive fund as the name of the parent directory
    fund = os.path.basename(os.path.dirname(fp))

    try:
        wb = load_workbook(fp, read_only=True, data_only=True)
        if "Book Capital" not in wb.sheetnames:
            return records


        ws = wb["Book Capital"]
        rows = ws.iter_rows(values_only=True)

        # 1) find the header row
        for row in rows:
            if row and row[0] == "K1 ID":
                headers = list(row)
                break
        else:
            # no K1 ID header in this sheet
            return records

        # 2) consume data rows until a blank row
        for row in rows:
            if all(cell is None for cell in row):
                break
            row_map = dict(zip(headers, row))
            k1 = row_map.get("K1 ID")
            ending = row_map.get("Ending Capital Balance")
            if k1 is not None:
                records.append({
                    "file_path": fp,
                    "Fund": fund,
                    "K1 ID": str(k1),
                    "Ending Capital Balance": ending
                })

    except Exception as e:
        # capture errors for debugging
        records.append({
            "file_path": fp,
            "Fund": fund,
            "K1 ID": None,
            "Ending Capital Balance": None,
            "error": str(e)
        })
    return records


def main():
    root = "folder/BOX_TAX_FOLDER"

    # 1) Gather all matching workbooks/workpapers, skipping any file containing 'BDO'
    files = []
    for dp, _, fns in os.walk(root):
        for fn in fns:
            lf = fn.lower()
            # ignore any file names that contain 'bdo'
            if 'bdo' in lf:
                continue
            if lf.endswith('.xlsx') and ('workbook' in lf or 'workpaper' in lf):
                files.append(os.path.join(dp, fn))

    # 2) Parallel scan with threads (no pickle issues)
    all_records = []
    with ThreadPoolExecutor(max_workers=os.cpu_count()) as pool:
        futures = [pool.submit(scan_book_capital, f) for f in files]
        for fut in as_completed(futures):
            all_records.extend(fut.result())

    # 3) Build DataFrame
    df = pd.DataFrame(all_records)

    # 4) Print the full DataFrame
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    print(df)

    # 5) Write out for external tools
    out_csv = "/Users/nickseta/Desktop/Ending Capital Balance/k1_ending_balances.csv"
    df.to_csv(out_csv, index=False)
    print(f"\nWrote full results to {out_csv} — you can now load that into Data Wrangler.")

if __name__ == "__main__":
    main()
